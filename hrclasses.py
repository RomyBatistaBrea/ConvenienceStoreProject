from openpyxl import load_workbook
import adminclasses
import pyautogui
import locale
locale.setlocale(locale.LC_ALL, '')


def check_employees_file():
    try:
        pyautogui.getWindowsWithTitle(
            'Excel')[0].activate()
        pyautogui.hotkey('alt', 'f4')
        pyautogui.sleep(1)
    except:
        pass
    pyautogui.press('winleft')
    pyautogui.sleep(0.4)
    pyautogui.write('excel')
    pyautogui.press('enter')
    pyautogui.sleep(1)  # wait for Excel to open

    # Get the dimensions of the screen
    screen_width, screen_height = pyautogui.size()

    # Set the position and size of the Excel application window
    excel_width = screen_width // 2
    excel_height = screen_height
    excel_x = screen_width - excel_width
    excel_y = 0
    pyautogui.getWindowsWithTitle('Excel')[0].activate()
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('enter')
    pyautogui.getWindowsWithTitle('Excel')[0].moveTo(excel_x, excel_y)
    pyautogui.getWindowsWithTitle(
        'Excel')[0].resizeTo(excel_width, excel_height)


class Employee:
    def __init__(self, employee_first_name, employee_last_name, age, email, pay_rate, workdays):
        self.employee_first_name = employee_first_name
        self.employee_last_name = employee_last_name
        self.age = age
        self.email = email
        self.workdays = workdays
        self.pay_rate = pay_rate


class ScheduleManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.employees = self.load_employees()

    def load_employees(self):
        try:
            wb = load_workbook(self.file_path)
        except FileNotFoundError:
            print(f"Error: Could not find file {self.file_path}")
            return []

        sheet = wb.active
        employees = []
        for row in sheet.iter_rows(values_only=True):
            employee = Employee(
                employee_first_name=row[0], employee_last_name=row[1], age=row[2], email=row[3], pay_rate=row[4], workdays=row[5])
            if str(employee.age) == "Age":
                continue
            else:
                employees.append(employee)  # Appends employees as objects
        return employees

    def get_employees(self):
        return self.employees

    def describe(self, employee_list):
        adminclasses.log_event(
            "Manager", "Viewed employee's personal information")
        choice = str(input("Enter employee email address: "))
        for employee in employee_list:
            if choice == employee.email:
                print(
                    f"Name: {employee.employee_first_name} {employee.employee_last_name}\nAge: {employee.age}\nEmail: {employee.email}\nPay rate: ${employee.pay_rate}/hour\nWork days: {employee.workdays}\n")
                break

    def schedule(self):  # finds employee and puts the days that they work
        adminclasses.log_event("Manager", "Scheduled an employee")
        try:
            wb = load_workbook(self.file_path)
        except FileNotFoundError:
            print(
                f"Error: Could not find file {self.file_path}")
            return []
        sheet = wb.active
        ws = sheet
        find = str(input('\nEnter employee email address: '))
        working = str(input('Enter days of work(M,T,W,TR,F,ST,SN): '))
        counter = 0
        for row in sheet.iter_rows(values_only=True):
            if find not in str(row[3]):
                counter += 1
            else:
                ws.cell(row=(counter+1), column=6).value = working
        try:
            wb.save(self.file_path)
        except:
            print("\nCould not save to file.")
        else:
            print(f"\nWork days have been added.")
        finally:
            check_employees_file()
            show_manager_menu()

    def remove_from_schedule(self):  # finds employee and puts "OFF"
        adminclasses.log_event(
            "Manager", "Removed an employee from the schedule")
        try:
            wb = load_workbook(self.file_path)
        except FileNotFoundError:
            print(
                f"Error: Could not find file {self.file_path}")
            return []
        sheet = wb.active
        ws = sheet
        find = str(input('\nEnter employee email address: '))
        off = "OFF"
        counter = 0
        for row in sheet.iter_rows(values_only=True):
            if find not in str(row[3]):
                counter += 1
            else:
                ws.cell(row=(counter+1), column=6).value = off
        try:
            wb.save(self.file_path)
        except:
            print("\nCould not save to file.")
        else:
            print(
                "\nIf employee exists, then employee is now off the schedule, otherwise, retry.")
        finally:
            check_employees_file()
            show_manager_menu()

    def payroll(self, email):
        try:
            wb = load_workbook(self.file_path)
        except FileNotFoundError:
            print(
                f"Error: Could not find file {self.file_path}")
            return []
        sheet = wb.active
        ws = sheet
        counter = 0
        total_payout = 0
        for row in sheet.iter_rows(values_only=True):
            if row[0] == "First Name":
                pass
            else:
                print(ws.cell(row=counter+1, column=8).value)
                print(ws.cell(row=counter+1, column=5).value)
                total_payout += (int(ws.cell(row=counter+1, column=8).value)
                                 ) * int(ws.cell(row=counter+1, column=5).value)
                ws.cell(row=counter+1, column=8).value = 0
            counter += 1
        print("\nTotal payout is " + locale.currency(total_payout, grouping=True))

        try:
            wb.save(self.file_path)
            adminclasses.log_event(
                "Manager", "Sent out and reset payroll")
        except FileNotFoundError:
            print("\nCould not save to file.")
        else:
            print("Payroll has been sent out and reset.")
        finally:
            check_employees_file()
            adminclasses.manager(email)

    def clock_in(self, email):
        try:
            wb = load_workbook(self.file_path)
        except FileNotFoundError:
            print(
                f"Error: Could not find file {self.file_path}")
            return []

        sheet = wb.active
        ws = sheet
        find = email
        counter = 0
        for row in sheet.iter_rows(values_only=True):
            if str(find) not in str(row[3]):
                counter += 1
            else:
                if str(ws.cell(row=(counter+1), column=7).value) == "1":
                    print("\nYou have already clocked in.")
                    return
                else:
                    ws.cell(row=(counter+1), column=7).value = "1"
        try:
            wb.save(self.file_path)
            check_employees_file()
        except:
            print("\nCould not save to file")
        else:
            print(f"\nYou have clocked in.")

    def clock_out(self, email):
        try:
            wb = load_workbook(self.file_path)
        except FileNotFoundError:
            print(
                f"Error: Could not find file {self.file_path}")
            return []

        sheet = wb.active
        ws = sheet
        find = email
        counter = 0
        for row in sheet.iter_rows(values_only=True):
            if str(find) not in str(row[3]):
                counter += 1
            else:
                x = int(ws.cell(row=(counter+1), column=8).value)
                if str(ws.cell(row=(counter+1), column=7).value) == "0":
                    print("\nYou have already clocked out.")
                    adminclasses.login()
                else:
                    ws.cell(row=(counter+1), column=7).value = "0"
                    ws.cell(row=(counter+1), column=8).value = x + 8
        try:
            wb.save(self.file_path)
            check_employees_file()
        except:
            print("\nCould not save to file")
        else:
            print(f"\nYou have clocked out.")

    def view_schedule(self, email):
        adminclasses.log_event("Manager", "Viewed employee schedule")
        try:
            wb = load_workbook(self.file_path)
        except FileNotFoundError:
            print(
                f"Error: Could not find file {self.file_path}")
            return []

        sheet = wb.active
        ws = sheet
        find = email
        counter = 0
        for row in sheet.iter_rows(values_only=True):
            if find not in str(row[3]):
                counter += 1
            else:
                print(f"\nSchedule: {row[5]}")

    def reset_password(self, email):
        adminclasses.log_event("Admin", "Reset an employee's password")
        try:
            wb = load_workbook(self.file_path)
        except FileNotFoundError:
            print(
                f"Error: Could not find file {self.file_path}")
            return []

        sheet = wb.active
        ws = sheet
        find = str(input('\nEnter employee email address: '))
        new_password = str(input('Enter new password: '))
        counter = 0
        try:
            for row in sheet.iter_rows(values_only=True):
                if find not in str(row[3]):
                    counter += 1
                else:
                    ws.cell(row=(counter+1), column=9).value = new_password
        except:
            print("\nEmployee not found.")
        else:

            try:
                wb.save(self.file_path)
            except:
                print("\nCould not save to file.")
            else:
                print(f"\nPassword has been changed.")
        finally:
            check_employees_file()
            adminclasses.admin(email)


def get_file_location():
    file_path = "C:/Users/romyb/Downloads/Employee_Info.xlsx"
    return ScheduleManager(file_path)


def show_manager_menu():
    choice = str(input(
        "\n1. Get Employee Schedule\n2. Edit Workdays for Employee\n3. Remove from Schedule\n4. Exit\n"))
    if choice == "1":
        ScheduleManager.describe(
            get_file_location(), ScheduleManager.get_employees(get_file_location()))
        show_manager_menu()
    elif choice == "2":
        ScheduleManager.schedule(get_file_location())
    elif choice == "3":
        ScheduleManager.remove_from_schedule(get_file_location())
    elif choice == "4":
        print("\nExiting...\n")
        exit()
    else:
        print("Invalid choice")
        ScheduleManager.show_manager_menu()
