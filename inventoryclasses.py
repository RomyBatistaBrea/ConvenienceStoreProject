# datetime to log orders
# random to generate order ID's for customers
# You have to install "openpyxl" to to run code.
# open powershell for windows type in "pip install openpyxl"
# restart computer
# Finance has to add tab for price and update class price and tax calcualtions

import datetime
import random
import openpyxl
import locale
import adminclasses
import pyautogui
locale.setlocale(locale.LC_ALL, '')

fp = '/Users/romyb/Downloads/Inventory.xlsx'


def check_fuel_file():
    try:
        pyautogui.getWindowsWithTitle(
            'Excel')[0].activate()
        pyautogui.hotkey('alt', 'f4')
        pyautogui.sleep(1)
    except:
        pass

    # Create an instance of the Excel application
    pyautogui.press('winleft')
    pyautogui.sleep(0.4)
    pyautogui.write('excel')
    pyautogui.press('enter')
    pyautogui.sleep(1.3)  # wait for Excel to open

    # Get the dimensions of the screen
    screen_width, screen_height = pyautogui.size()

    # Set the position and size of the Excel application window
    excel_width = screen_width // 2
    excel_height = screen_height
    excel_x = screen_width - excel_width
    excel_y = 0
    pyautogui.getWindowsWithTitle('Excel')[0].activate()
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('down')
    pyautogui.press('down')
    pyautogui.press('enter')
    pyautogui.getWindowsWithTitle('Excel')[0].moveTo(excel_x, excel_y)
    pyautogui.getWindowsWithTitle(
        'Excel')[0].resizeTo(excel_width, excel_height)


def check_inventory_file():
    try:
        pyautogui.getWindowsWithTitle(
            'Excel')[0].activate()
        pyautogui.hotkey('alt', 'f4')
        pyautogui.sleep(1)
    except:
        pass

    # Create an instance of the Excel application
    pyautogui.press('winleft')
    pyautogui.sleep(0.4)
    pyautogui.write('excel')
    pyautogui.press('enter')
    pyautogui.sleep(1.3)  # wait for Excel to open

    # Get the dimensions of the screen
    screen_width, screen_height = pyautogui.size()

    # Set the position and size of the Excel application window
    excel_width = screen_width // 2
    excel_height = screen_height
    excel_x = screen_width - excel_width
    excel_y = 0
    pyautogui.getWindowsWithTitle('Excel')[0].activate()
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('down')
    pyautogui.press('down')
    pyautogui.press('down')
    pyautogui.press('enter')
    pyautogui.getWindowsWithTitle('Excel')[0].moveTo(excel_x, excel_y)
    pyautogui.getWindowsWithTitle(
        'Excel')[0].resizeTo(excel_width, excel_height)


class Item:
    def __init__(self,  item_type, item_id, item_name, quantity, price):
        self.item_type = item_type
        self.item_id = item_id
        self.item_name = item_name
        self.quantity = quantity
        self.price = price


class Cart:
    def __init__(self, cart):
        self.cart = cart


class InventoryManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.items = self.load_items()

    def load_items(self):
        try:
            wb = openpyxl.load_workbook(self.file_path)
        except FileNotFoundError:
            print(f"Error: Could not find file {self.file_path}")
            return []

        sheet = wb.active
        items = []
        for row in sheet.iter_rows(values_only=True):
            item = Item(item_id=row[1], item_name=row[2],
                        item_type=row[0], quantity=row[3],
                        price=row[4])
            if item.quantity == "Quantity":
                continue
            else:
                items.append(item)  # Appends items as objects
        return items

    def update_items(self, items):
        max_item_capacity = 0
        try:
            wb = openpyxl.load_workbook(self.file_path)
        except FileNotFoundError:
            print(f"Error: Could not find file {self.file_path}")
            return

        sheet = wb.active
        ws = sheet
        counter = 0
        for row in sheet.iter_rows(values_only=True):
            if str(row[1]) != 50:
                ws.cell(row=(counter+1),
                        column=5).value = max_item_capacity
                wb.save(self.file_path)

    def get_cart(self, new_price):
        option = ""
        x = True
        while x:
            option = input("Enter item: ")
            for item in self.load_items():
                try:
                    if int(option) == int(item.item_id):
                        try:
                            wb = openpyxl.load_workbook(self.file_path)
                        except FileNotFoundError:
                            print(
                                f"Error: Could not find file {self.file_path}")
                            return []
                        sheet = wb.active
                        ws = sheet
                        counter = 0
                        for row in sheet.iter_rows(values_only=True):
                            if option not in str(row[1]):
                                counter += 1
                            else:
                                ws.cell(row=(counter+1),
                                        column=5).value = new_price
                                try:
                                    wb.save(self.file_path)
                                except:
                                    print("\nInvalid input")
                                finally:
                                    show_inventory_menu()
                                x = False
                                break
                except Exception as e:
                    x = False
                    print(f"Error: {e}")

    def restock(self, updated_quantity):
        file_path = self.file_path
        try:
            wb = openpyxl.load_workbook(
                '/Users/romyb/Downloads/Inventory.xlsx')
        except FileNotFoundError:
            print(f"Error: Could not find file {file_path}")
            return

        sheet = wb.active
        sheet.delete_rows(2, sheet.max_row)  # clear existing items
        counter = 1
        for items in updated_quantity:
            sheet.append([items[0], items[1], items[2], items[3], items[4]])
            sheet.cell(row=counter+1, column=2).alignment = openpyxl.styles.Alignment(
                horizontal='right')
            counter += 1
        wb.save(file_path)

    def get_inventory(self):
        return self.items

    def get_low_inventory(self):
        low_inventory_items = []
        threshold = 50.0

        for item in self.items:
            if item.quantity < threshold:
                low_inventory_items.append(item)

        if len(low_inventory_items) == 0:
            low_inventory_items = None
            print("No items need to be restocked.")
        elif len(low_inventory_items) > 0:
            return low_inventory_items

    def place_order(self):
        items = self.get_inventory()
        updated_items = []
        total_price = 0
        counter = 0

        for item in items:
            price = str(item.price).replace(",", "")
            price = price.replace("$", "")
            price = float(price)
            quantity = int(item.quantity)

            if quantity <= 50:
                total_price += price * (50 - quantity)
                if quantity < 50:
                    counter += 1

            updated_items.append([
                item.item_type, item.item_id, item.item_name, 50, locale.currency(float(price), grouping=True)])
        if counter == 0:
            adminclasses.log_event(
                "Inventory", "Viewed Inventory")
            print("\nAll items are stocked, none need to be ordered.")
        else:
            adminclasses.log_event(
                "Inventory", "Viewed and Restocked Inventory")
            self.restock(updated_items)
            print(
                f"\nAll items have been restocked to 50 units and logged in the order log.\nTotal price: {locale.currency(float(total_price), grouping=True)}")
            # print("Bank: locale.currency(float(BANK), grouping=True)")
            order_id = self.generate_order_id()
            log_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            order_log = f"Order ID: {order_id}, Total Price: {locale.currency(total_price, grouping=True)}, Time: {log_time}"
            pyautogui.getWindowsWithTitle('Excel')[0].activate()
            pyautogui.hotkey('alt', 'f4')
            check_inventory_file()
            self.log_order(order_log)

    def create_item(self):  # Product creation - add details about product

        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        category = input(
            ('\nPlease enter the product category (Snacks, Drinks, or Food)\n----------------------------------\n'))
        while category.capitalize() not in ['Snacks', 'Drinks', 'Food']:
            print(
                'Invalid category. Please try again.\n----------------------------------\n')
            category = input(
                ('\nPlease enter the product category (Snacks, Drinks, or Food)\n----------------------------------\n'))
        id = self.generate_order_id()
        for row in ws.iter_rows(values_only=True):
            while id in row:
                id = self.generate_order_id()
        name = input(
            ('Please enter the product name\n----------------------------------\n'))
        for row in ws.iter_rows(values_only=True):
            while name in row:
                print(
                    'Item already exists. Please try again.\n----------------------------------\n')
                name = input(
                    ('Please enter the product name\n----------------------------------\n'))
        quantity = 0
        price = float(
            input(('Please enter the product price\n----------------------------------\n')))
        item = Item(category, id, name, quantity, price)
        return item

    # Completes the process and registers a new item for inventory and saves it in excel
    def register_item(self, item):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        confirmation = input(
            (f'\nProduct Entered-\nCategory: {item.item_type.capitalize()}\nName: {item.item_name.capitalize()}\nUnique ID: {int(item.item_id)}\nPrice: {locale.currency(item.price, grouping=True)}\nDoes this look good? (1. Yes / 2. No)\n----------------------------------\n'))
        if confirmation == '1':
            ws.cell(row=(ws.max_row+1),
                    column=1).value = item.item_type.capitalize()
            ws.cell(row=(ws.max_row), column=2).value = item.item_id
            ws.cell(row=(ws.max_row), column=2).alignment = openpyxl.styles.Alignment(
                horizontal='right')
            ws.cell(row=(ws.max_row), column=3).value = item.item_name.capitalize()
            ws.cell(row=(ws.max_row), column=4).value = 0
            ws.cell(row=(ws.max_row), column=5).value = (
                locale.currency(item.price, grouping=True))
            wb.save(self.file_path)
            pyautogui.getWindowsWithTitle('Excel')[0].activate()
            pyautogui.hotkey('alt', 'f4')
            check_inventory_file()
            print('Item registered!\n')
        elif item == '2':
            print('Item registration cancelled\n----------------------------------\n')
            show_inventory_menu()

    def item_search(self):  # Allows the customer to search for products by product id and name
        wb = openpyxl.load_workbook(fp)
        ws = wb.active
        search = input(
            'Please enter product ID or search term. Type "1" when done searching\n')
        truth = False
        while str(search) != str(1):
            for row in ws.iter_rows(values_only=True):
                if search in str(row):
                    truth = True
                    print(f'{row}\n----------------------------------\n')
            if truth == False:
                print('No items found!')
            search = input(
                'Please enter product ID or search term. Type "1" when done searching\n')
        if str(search) == str(1):
            print('Search finished!')

    def create_cart(self):  # Creates a cart for the customer
        wb = openpyxl.load_workbook(fp)
        ws = wb.active
        cart = []
        while True:
            try:
                cart_creation = int(input(
                    'Please enter the ID of a product you would like to add to your cart. Enter 1 when finished\n'))
                break
            except ValueError:
                print('Must enter numbers!')
                continue
        while int(cart_creation) != int(1):
            if int(cart_creation) in range(111111, 999999):
                for row in ws.iter_rows(values_only=True):
                    if str(cart_creation) in str(row):
                        cart.append(cart_creation)
            else:
                print('Invalid ID')
            while True:
                try:
                    cart_creation = int(input(
                        'Please enter the ID of a product you would like to add to your cart. Enter 1 when finished\n'))
                    break
                except ValueError:
                    print('Must enter numbers!')
                    continue
        print(
            f'Your cart contains {cart}\n----------------------------------\n')
        newcart = Cart(cart)
        return newcart

    # Fully processes the customer's purchase and subtracts from our inventory
    def purchase_cart(self, newcart):
        try:
            wb = openpyxl.load_workbook(
                '/Users/romyb/Downloads/Inventory.xlsx')
        except FileNotFoundError:
            print(f"Error: Could not find file {fp}")
            return
        sheet = wb.active
        ws = sheet
        quantity = []
        ordertotal = 0

        # Asking user to enter quantity. Checking stock and that quantity is int. Appending to list.
        for id in newcart.cart:
            count3 = 0
            for row in ws.iter_rows(values_only=True):
                if str(id) not in str(row):
                    count3 += 1
                elif str(id) in str(row):
                    print(f'Cart Item: {row}\n')
                    while True:
                        try:
                            choose_quantity = int(
                                input('Quantity of item being bought?\n----------------------------\n'))
                            break
                        except ValueError:
                            print('You must enter a number!')
                            continue
                    while int(choose_quantity) > int(ws.cell(row=(count3+1), column=4).value):
                        print('\nQuantity chosen exceeds stock\n')
                        while True:
                            try:
                                choose_quantity = int(
                                    input('Quantity of item being bought?\n----------------------------\n'))
                                break
                            except ValueError:
                                print('You must enter a number!')
                                continue
                    quantity.append((id, choose_quantity))

        # calculate order cost, subtract quantity from inventory
        for id, choose_quantity in quantity:
            count = 0
            for row in ws.iter_rows(values_only=True):
                if str(id) not in str(row):
                    count += 1
                elif str(id) in str(row):
                    itemcost = float(
                        str(ws.cell(row=(count + 1), column=5).value).replace('$', ''))
                    ordertotal += (itemcost * choose_quantity)
        taxedtotal = float(ordertotal * 1.06)
        orderconfirmation = input(
            f'Order total is {locale.currency(taxedtotal, grouping=True)} including tax. Proceed to purchase? (1. Yes, 2. No)\n')
        while orderconfirmation not in ('1', '2'):
            print('Please enter 1 or 2!')
            orderconfirmation = input(
                f'Order total is {locale.currency(taxedtotal, grouping=True)} including tax. Proceed to purchase? (1. Yes, 2. No)\n')
        if orderconfirmation == '1':
            # add function to add money to bank file
            for id, choose_quantity in quantity:
                count2 = 0
                for row in ws.iter_rows(values_only=True):
                    if str(id) not in str(row):
                        count2 += 1
                    elif str(id) in str(row):
                        currentquantity = int(
                            ws.cell(row=(count2 + 1), column=4).value)
                        newquantity = ((currentquantity) -
                                       int(choose_quantity))
                        ws.cell(row=(count2 + 1),
                                column=4).value = str(newquantity)
                        ws.cell(row=(
                            count2 + 1), column=4).alignment = openpyxl.styles.Alignment(horizontal='right')
            wb.save(fp)
            print('Thank you for your purchase!')
            check_inventory_file()
        elif orderconfirmation == '2':
            print('Order cancelled')
            return

    # Creates a random id for products/order id's/event logger
    def generate_order_id(self):
        return str(random.randint(100000, 999999))

    def log_order(self, order_log):  # Order logger, different to the event logger for admin
        try:
            with open("order.log", "a") as f:
                f.write(order_log + "\n")
        except Exception as e:
            print(f"Error writing to order log: {e}")


class Order:
    def __init__(self, quantity):
        self.quantity = quantity

    def calculate_total_price(self, item_list):
        if item_list is None:
            return 0
        else:
            total = 0
            for item in item_list:
                price = str(item.price).replace(",", "")
                price = price.replace("$", "")
                price = float(price)
                total += price * (50 - int(item.quantity))
            return locale.currency(total, grouping=True)


def get_file_location():
    return InventoryManager(
        '/Users/romyb/Downloads/Inventory.xlsx')


def correct_price():
    adminclasses.log_event("Inventory", "Updated item Price")
    option = ""
    final_say = 0  # checks if the id exists in the inventory sheet or not
    option = input("\nEnter the id of the item: ")
    for item in InventoryManager.load_items(get_file_location()):
        try:
            if int(option) == int(item.item_id):
                try:
                    wb = openpyxl.load_workbook(
                        get_file_location().file_path)
                except FileNotFoundError:
                    print(
                        f"Error: Could not find file {get_file_location().file_path}")
                sheet = wb.active
                ws = sheet
                counter = 0
                for row in sheet.iter_rows(values_only=True):
                    if option not in str(row[1]):
                        counter += 1
                    elif option in str(row[1]):
                        price = str(row[4]).replace(",", "")
                        price = price.replace("$", "")
                        price = float(price)
                        print(
                            f"{row[2].upper()}, {locale.currency(price, grouping=True)}")
                        q = counter
                        choose = input("Is this the correct item? (Y/N): ")
                        if choose.upper() == "Y":
                            y = True
                            while y:
                                try:
                                    new_price = float(
                                        input("Enter the new price: $"))
                                    while new_price < float(0.01):
                                        try:
                                            if new_price < float(0.01):
                                                print(
                                                    "\nPrice has to be greater than 0")
                                            new_price = float(
                                                input("\nEnter the new price: "))
                                        except ValueError:
                                            print(
                                                "\nYou must enter numbers")
                                    y = False
                                except ValueError:
                                    print("\nYou must enter numbers")
                                else:
                                    print(
                                        f"\nPrice has been updated to {locale.currency(new_price, grouping=True)}")

                        elif choose.upper() == "N":
                            correct_price()
                        else:
                            print("\nInvalid input")
                            correct_price()

                ws.cell(row=(q+1),
                        column=5).value = locale.currency(new_price, grouping=True)

                try:
                    final_say = 1
                    wb.save(get_file_location().file_path)
                    pyautogui.getWindowsWithTitle(
                        'Excel')[0].activate()
                    pyautogui.hotkey('alt', 'f4')
                    check_inventory_file()
                except:
                    print("\nCould not find file")
                finally:
                    show_inventory_menu()

        except ValueError:
            print(f"\nInvalid id, try again")
            correct_price()

    if final_say == 0:
        print("\nInvalid id, try again.")
        correct_price()


def show_inventory_menu():
    choice = str(input("\n1. Update item price\n2. Check Current Inventory\n3. Get Restockable Items list\n4. Calculate Total for Low Inventory Order\n5. Register a New Item\n6. Exit\n"))
    if choice == "1":
        correct_price()
    elif choice == "2":
        for item in InventoryManager.get_inventory(get_file_location()):
            price = str(item.price).replace(",", "")
            price = price.replace("$", "")
            price = float(price)
            print(
                f"{item.item_type}, {int(item.item_id)}, {item.item_name}, Quantity: {int(item.quantity)}, Price: {locale.currency(price, grouping=True)}")
        show_inventory_menu()
    elif choice == "3":
        print("\n")
        if InventoryManager.get_low_inventory(get_file_location()) is None:
            show_inventory_menu()
        else:
            for item in InventoryManager.get_low_inventory(get_file_location()):
                print(f"{item.item_name}: {int(item.quantity)} units left")
            show_inventory_menu()
    elif choice == "4":
        print("")
        x = Order.calculate_total_price(
            InventoryManager.load_items(get_file_location()), InventoryManager.get_low_inventory(get_file_location()))
        if x == 0:
            print("")
        else:
            print(f"\nTotal price for restock order: {x}")
        show_inventory_menu()
    elif choice == "5":
        InventoryManager.register_item(
            get_file_location(), item=InventoryManager.create_item(get_file_location()))
        show_inventory_menu()
    elif choice == "6":
        InventoryManager.place_order(get_file_location())
        print("\nHave a great day! Closing Now...")
        exit()
    else:
        print("Invalid")
        show_inventory_menu()
