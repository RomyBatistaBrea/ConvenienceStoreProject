# admin classes
import openpyxl
import datetime
import inventoryclasses
import hrclasses
import pyautogui

file_path = '/Users/romyb/Downloads/Employee_Info.xlsx'
today = datetime.date.today()
d1 = today.strftime("%m/%d/%Y")


def check_log_file():
    # Create an instance of the Excel application
    # Create an instance of the Excel application
    pyautogui.press('winleft')
    pyautogui.write('excel')
    pyautogui.press('enter')
    pyautogui.sleep(1.3)  # wait for Excel to open

    # Get the dimensions of the screen
    screen_width, screen_height = pyautogui.size()

    # Set the position and size of the Excel application window
    excel_width = screen_width // 2
    excel_height = screen_height
    excel_x = screen_width - excel_width
    excel_y = 0
    pyautogui.getWindowsWithTitle('Excel')[0].activate()
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('down')
    pyautogui.press('enter')
    pyautogui.getWindowsWithTitle('Excel')[0].moveTo(excel_x, excel_y)
    pyautogui.getWindowsWithTitle(
        'Excel')[0].resizeTo(excel_width, excel_height)


def check_event_logs_file():
    # Create an instance of the Excel application
    # Create an instance of the Excel application
    pyautogui.press('winleft')
    pyautogui.write('excel')
    pyautogui.press('enter')
    pyautogui.sleep(1.3)  # wait for Excel to open

    # Get the dimensions of the screen
    screen_width, screen_height = pyautogui.size()

    # Set the position and size of the Excel application window
    excel_width = screen_width // 2
    excel_height = screen_height
    excel_x = screen_width - excel_width
    excel_y = 0
    pyautogui.getWindowsWithTitle('Excel')[0].activate()
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('down')
    pyautogui.press('enter')
    pyautogui.getWindowsWithTitle('Excel')[0].moveTo(excel_x, excel_y)
    pyautogui.getWindowsWithTitle(
        'Excel')[0].resizeTo(excel_width, excel_height)


class Login:

    def __init__(self, email_address, password, position, hire_date):
        self.email_address = email_address
        self.password = password
        self.position = position
        self.hire_date = hire_date


class Log:
    def __init__(self, date, event_type, message):
        self.date = date
        self.event_type = event_type
        self.message = message

# -------------------------------------------------------------------------------------------------------------------------
# EVENT LOGGER FUNCTION STARTS HERE
# -------------------------------------------------------------------------------------------------------------------------


def load_logs(log_path):
    try:
        wb = openpyxl.load_workbook(log_path)
    except FileNotFoundError:
        print(f"Error: Could not find file {log_path}")
        return []

    sheet = wb.active
    logs = []
    for row in sheet.iter_rows(values_only=True):
        log = Log(date=row[0], event_type=row[1], message=row[2])

        logs.append(log)  # Appends logs as objects
    return logs


def log_event(event_type, message):
    log_path = "C:/Users/romyb/Downloads/Event_Logs.xlsx"
    try:
        wb = openpyxl.load_workbook(log_path)
    except FileNotFoundError:
        print(f"Error: Could not find file {log_path}")
        return []

    sheet = wb.active
    ws = sheet
    ws.cell(row=(ws.max_row+1), column=1).value = d1
    ws.cell(row=(ws.max_row), column=2).value = event_type
    ws.cell(row=(ws.max_row), column=3).value = message
    try:
        wb.save(log_path)
    except:
        print("\nCould not save to file.")
    else:
        return []

# -------------------------------------------------------------------------------------------------------------------------
# EMPLOYEE LOADER FUNCTION STARTS HERE
# -------------------------------------------------------------------------------------------------------------------------


def load_people(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"Error: Could not find file {file_path}")
        return []

    sheet = wb.active
    people = []
    for row in sheet.iter_rows(values_only=True):
        person = Login(email_address=row[3], password=row[8],
                       position=row[9], hire_date=row[10])

        people.append(person)  # Appends people as objects
    return people

# -------------------------------------------------------------------------------------------------------------------------
# ADMIN FUNCTIONS START HERE
# -------------------------------------------------------------------------------------------------------------------------


def admin(email_address):
    choice = input(
        "\n\nWelcome, what would you like to do?\n1. Clock in\n2. Clock out\n3. View Schedule\n4. Reset password\n5. Log Management\n6. Exit\n")

    if choice == "1":
        hrclasses.ScheduleManager.clock_in(
            hrclasses.get_file_location(), email_address)
        admin(email_address)
    elif choice == "2":
        hrclasses.ScheduleManager.clock_out(
            hrclasses.get_file_location(), email_address)

    elif choice == "3":
        hrclasses.ScheduleManager.view_schedule(
            hrclasses.get_file_location(), email_address)
        admin(email_address)

    elif choice == "4":
        hrclasses.ScheduleManager.reset_password(
            hrclasses.get_file_location(), email_address)
        admin(email_address)

    elif choice == "5":
        choice = input("\nLog Management\n1. View log\n2. Back\n")
        if choice == "1":
            check_event_logs_file()
            for row in load_logs("C:/Users/romyb/Downloads/Event_Logs.xlsx"):
                print(f"{row.date} {row.event_type} {row.message}")
            admin(email_address)

        elif choice == "2":
            admin(email_address)
        else:
            print("Invalid choice")
            admin(email_address)

    elif choice == "6":
        print("Exiting....")
        exit()
    else:
        print("Invalid choice")
        admin(email_address)

# -------------------------------------------------------------------------------------------------------------------------
# MANAGER FUNCTIONS START HERE
# -------------------------------------------------------------------------------------------------------------------------


def reinput_manager():
    position = input(
        "\nEmployee position:\n1. Admin\n2. Manager\n3. Employee\n")
    if position == "1":
        x = "admin"
    elif position == "2":
        x = "manager"
    elif position == "3":
        x = "employee"
    else:
        print("Invalid choice")
        reinput_manager()
    return x


def manager(email_address):
    choice = input(
        "\n\nWhat would you like to do?\n1. Clock in\n2. Clock out\n3. View my schedule\n4. Employee management\n5. Log management\n6. Inventory management\n7. Log out\n")

    if choice == "1":
        hrclasses.ScheduleManager.clock_in(
            hrclasses.get_file_location(), email_address)
        manager(email_address)
    elif choice == "2":
        hrclasses.ScheduleManager.clock_out(
            hrclasses.get_file_location(), email_address)
    elif choice == "3":
        hrclasses.ScheduleManager.view_schedule(
            hrclasses.get_file_location(), email_address)
        manager(email_address)
    elif choice == "4":
        hrclasses.check_employees_file()
        choice = input(
            "\nEmployee management\n1. Hire\n2. Fire\n3. Employees Schedule Management\n4. Send Out Payroll\n5. Back\n")
        if choice == "1":
            first_name = input("Enter new employee first name: ")
            last_name = input("Enter new employee last name: ")
            age = input("Enter new employee age: ")
            pay = input("Enter new employee pay: ")
            email_address = input(
                "----------------------------------------------------------------\nHire - New employee email address: ")
            password = input("Enter new employee password: ")
            pos = reinput_manager()

            try:
                wb = openpyxl.load_workbook(file_path)
            except FileNotFoundError:
                print(f"Error: Could not find file {file_path}")
                return

            sheet = wb.active
            ws = sheet

            try:
                ws.cell(row=(ws.max_row+1), column=1).value = str(first_name)
                ws.cell(row=(ws.max_row), column=2).value = str(last_name)
                ws.cell(row=(ws.max_row), column=3).value = str(age)
                ws.cell(row=(ws.max_row), column=4).value = str(email_address)
                ws.cell(row=(ws.max_row), column=5).value = str(pay)
                ws.cell(row=(ws.max_row), column=7).value = 0
                ws.cell(row=(ws.max_row), column=8).value = 0
                ws.cell(row=(ws.max_row), column=9).value = password
                ws.cell(row=(ws.max_row), column=10).value = str(pos)
                ws.cell(row=(ws.max_row), column=11).value = d1
                ws.cell(row=(ws.max_row), column=12).value = "1"
                wb.save(file_path)
                hrclasses.check_employees_file()
                print("New employee added! Closing program...")
            except:
                print("Error: Could not add new employee")
                manager(email_address)

        elif choice == "2":
            employee = input(
                "----------------------------------------------------------------\nFire - Enter employee email address: ")
            try:
                wb = openpyxl.load_workbook(file_path)
            except FileNotFoundError:
                print(f"Error: Could not find file {file_path}")
                return

            sheet = wb.active
            ws = sheet
            counter = 0
            try:
                for row in sheet.iter_rows(values_only=True):
                    counter += 1
                    if row[3] == employee:
                        ws.cell(row=(counter), column=6).value = "OFF"
                        ws.cell(row=(counter), column=12).value = "0"
                        wb.save(file_path)
                        hrclasses.check_employees_file()
                        print("Employee flagged - Inactive!\nClosing program...")
                        break
            except:
                print("\nError: Could not remove employee\n")
                manager(email_address)
        elif choice == "3":
            hrclasses.show_manager_menu()
        elif choice == "4":
            hrclasses.ScheduleManager.payroll(
                hrclasses.get_file_location(), email_address)
        elif choice == "5":
            manager(email_address)
        else:
            print("Invalid choice")
            manager(email_address)

    elif choice == "5":
        choice = input("\nLog management\n1. View log\n2. Back\n")
        if choice == "1":
            print("\nView order log")
            with open("order.log", "r") as file:
                print(file.read())
                manager(email_address)
        elif choice == "2":
            manager(email_address)
        else:
            print("Invalid choice")
            manager(email_address)

    elif choice == "6":
        inventoryclasses.check_inventory_file()
        inventoryclasses.show_inventory_menu()
    elif choice == "7":
        print("Logging out...")
        exit()
    else:
        print("Invalid choice")
        manager(email_address)

# -------------------------------------------------------------------------------------------------------------------------
# EMPLOYEE FUNCTIONS START HERE
# -------------------------------------------------------------------------------------------------------------------------


def employee(email_address):
    choice = input(
        "\n\nWelcome, what would you like to do?\n1. Clock in\n2. Clock out\n3. View schedule\n4. Log out\n")
    if choice == "1":

        hrclasses.ScheduleManager.clock_in(
            hrclasses.get_file_location(), email_address)
        employee(email_address)
    elif choice == "2":
        hrclasses.ScheduleManager.clock_out(
            hrclasses.get_file_location(), email_address)
    elif choice == "3":
        hrclasses.ScheduleManager.view_schedule(
            hrclasses.get_file_location(), email_address)
        employee(email_address)
    elif choice == "4":
        print("Logging out...")
        exit()
    else:
        print("\nInvalid choice")
        employee(email_address)

# -------------------------------------------------------------------------------------------------------------------------
# Customer function starts here
# -------------------------------------------------------------------------------------------------------------------------


def customer():
    choose = input("\n1. Buy fuel\n2. Purchase items\n3. Exit\n")
    if choose == '1':
        inventoryclasses.check_fuel_file()

        pass
    elif choose == '2':
        inventoryclasses.check_inventory_file()
        search_or_buy = input(
            '\n1. Search the catalogue\n2. Add items to cart\n----------------------------\n')
        if search_or_buy == '1':
            inventoryclasses.InventoryManager.item_search(1)
            searchend = input(
                '\nWould you like to begin adding items to the cart? (1. Yes 2. No)\n-------------------------------\n')
            if searchend == '1':
                inventoryclasses.InventoryManager.purchase_cart(
                    1, newcart=inventoryclasses.InventoryManager.create_cart(1))
            elif searchend == '2':
                print("Have a nice day!")
            else:
                print('Invalid option!')
        elif search_or_buy == '2':
            inventoryclasses.InventoryManager.purchase_cart(
                1, newcart=inventoryclasses.InventoryManager.create_cart(1))
        else:
            print('Invalid Option')
    elif choose == '3':
        print('Have a nice day!')
        exit()
    else:
        print('Invalid option!')

# -------------------------------------------------------------------------------------------------------------------------
# THE MAIN FUNCTION TO CHECK EMPLOYEE POSITION STARTS HERE
# LOGIN CHECK
# -------------------------------------------------------------------------------------------------------------------------


def login():
    username = str(input("\nEnter username: "))
    password = str(input("Enter password: "))

    truth = False
    for person in load_people(file_path):
        if str(username) == str(person.email_address) and str(password) == str(person.password):
            if str(person.position) == "admin":
                truth = True
                admin(person.email_address)
            elif str(person.position) == "manager":
                truth = True
                manager(person.email_address)
            elif str(person.position) == "employee":
                truth = True
                employee(person.email_address)
            else:
                break
    if truth == False:
        print("\nInvalid username or password\n")
        login()
